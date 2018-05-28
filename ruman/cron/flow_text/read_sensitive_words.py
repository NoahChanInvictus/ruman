# -*- coding:utf-8 -*-

from openpyxl import load_workbook
import redis
import json
import sys
reload(sys)
sys.path.append('../../')
from global_utils import R_ADMIN as r
#r = redis.StrictRedis(host="10.128.55.69", port="6379", db=15)
data = load_workbook('sensitive_words.xlsx')
table = data.get_sheet_by_name('Sheet2')
for i in range(1,549):
    word = table.cell(row=i, column=0).value
    level = table.cell(row=i, column=1).value
    r.hset('sensitive_words',word, level)
print r.hkeys('sensitive_words')




